"""
Ingestion + cleaning layer.

Handles the fact that real-world exports (MS Project / PMO tool exports, in
this case) are messy: inconsistent column names between projects, broken
formula cells ("#UNPARSEABLE"), blank predecessor/description fields that are
normal, and duplicate/derived columns. This module is intentionally
deterministic (no LLM) — data hygiene should never depend on model judgement.

This module only ever deals with a *local file path* — it has no idea
whether that path came from `sample_data/` or was just synced down from a
Google Drive folder a minute ago (see ingestion/source.py and
ingestion/drive_client.py). That seam is what lets the data source change
in production (local -> Drive -> some future PMO API) without touching this
file, metrics.py, llm_client.py, or the graph.
"""
from __future__ import annotations

import re

import openpyxl
import pandas as pd

from project_health_agent.core.exceptions import PlanParseError
from project_health_agent.core.logging_config import get_logger

logger = get_logger("ingestion.data_loader")

UNPARSEABLE = "#UNPARSEABLE"

# Canonical column name -> list of aliases seen across project plan exports.
# New project templates just need an alias added here; nothing else changes.
COLUMN_ALIASES = {
    "task_name": ["Task Name"],
    "status": ["Status"],
    "pct_complete": ["% Complete"],
    "start_date": ["Start Date"],
    "end_date": ["End Date"],
    "priority": ["Priority"],
    "owner": ["Owner"],
    "on_hold": ["On Hold?"],
    "not_applicable": ["Not Applicable?"],
    "duration": ["Duration"],
    "predecessors": ["Predecessors"],
    "total_float": ["Total Float"],
    "critical": ["Critical ?", "Critical?"],
    "baseline_start": ["Baseline Start"],
    "baseline_finish": ["Baseline Finish"],
    "variance": ["Variance"],
    "status_comment": ["Status Comment"],
    "comments": ["Comments"],
    "assigned_to": ["Assigned To"],
    "phase_milestone": ["Phase/Milestone"],
    "area": ["Area"],
    "at_risk": ["At Risk?"],
    "schedule_health": ["Schedule Health"],
    "project_manager": ["Project Manager"],
    "ancestors": ["Ancestors"],
    "task_rag": ["RAG"],
    "description": ["Description"],
}

# Columns that must be present and non-blank for a task row to count toward
# the data-completeness signal. Predecessors/Description/Comments are
# legitimately blank for most tasks, so they're excluded from this check.
CORE_SIGNAL_COLUMNS = [
    "task_name", "status", "pct_complete", "start_date", "end_date",
    "duration", "baseline_start", "baseline_finish", "variance",
]


def _build_rename_map(columns):
    rename = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in columns:
                rename[alias] = canonical
                break
    return rename


def _parse_duration_days(value):
    """'170d' -> 170.0, '0' -> 0.0, None/blank -> None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.match(r"^-?\d+(\.\d+)?", str(value).strip())
    return float(match.group()) if match else None


def _parse_variance_days(value):
    """'-2d' -> -2.0, '0' -> 0.0."""
    return _parse_duration_days(value)


def _clean_value(v):
    if v == UNPARSEABLE or v == "":
        return None
    return v


def _find_task_sheet_name(sheet_names):
    for name in sheet_names:
        if name not in ("Comments", "Summary"):
            return name
    raise PlanParseError("Could not find a task sheet (only Comments/Summary present).")


def load_project_plan(path: str) -> dict:
    """
    Load and clean a project plan workbook.

    Returns a dict:
        {
          "project_name": str,
          "summary": {...cleaned Summary sheet as dict...},
          "tasks": pd.DataFrame  # normalized, cleaned task-level rows
          "comments_log": [ {row, comment, author, timestamp}, ... ],
          "data_completeness": float in [0, 1],
          "source_path": str,
        }
    """
    logger.info("Loading project plan: %s", path)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except (OSError, KeyError) as exc:
        raise PlanParseError(f"Could not open workbook '{path}': {exc}") from exc

    task_sheet_name = _find_task_sheet_name(wb.sheetnames)
    ws = wb[task_sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise PlanParseError(f"Task sheet '{task_sheet_name}' in '{path}' is empty.")
    header, data_rows = rows[0], rows[1:]
    df = pd.DataFrame(data_rows, columns=header)
    df = df.rename(columns=_build_rename_map(list(header)))

    # Keep only canonical columns we recognize (extras like duplicate
    # Baseline Start2/Variance2/derived '#UNPARSEABLE' helper columns from
    # older templates are dropped rather than silently misread).
    keep_cols = [c for c in COLUMN_ALIASES.keys() if c in df.columns]
    df = df[keep_cols].copy()

    # Clean #UNPARSEABLE / blanks -> None
    for col in df.columns:
        df[col] = df[col].apply(_clean_value)

    if "duration" in df.columns:
        df["duration_days"] = df["duration"].apply(_parse_duration_days)
    if "variance" in df.columns:
        df["variance_days"] = df["variance"].apply(_parse_variance_days)
    if "critical" in df.columns:
        df["critical"] = df["critical"].apply(lambda v: bool(v) if v is not None else False)
    if "on_hold" in df.columns:
        df["on_hold"] = df["on_hold"].apply(lambda v: bool(v) if v is not None else False)
    if "not_applicable" in df.columns:
        df["not_applicable"] = df["not_applicable"].apply(lambda v: bool(v) if v is not None else False)

    # Project name = the root task row (Ancestors == 0), falling back to the
    # first row's task_name if the Ancestors column isn't present.
    project_name = None
    if "ancestors" in df.columns:
        root_rows = df[df["ancestors"] == 0]
        if not root_rows.empty:
            project_name = root_rows.iloc[0]["task_name"]
    if project_name is None and not df.empty:
        project_name = df.iloc[0]["task_name"]

    # --- Summary sheet ---
    summary = {}
    if "Summary" in wb.sheetnames:
        for key, val in wb["Summary"].iter_rows(values_only=True):
            if key is not None:
                summary[key] = _clean_value(val)

    # --- Comments sheet (free-form PM/client log, separate from per-task
    #     Status Comment / Comments columns) ---
    comments_log = []
    if "Comments" in wb.sheetnames:
        for row in wb["Comments"].iter_rows(values_only=True):
            if row and row[0] and str(row[0]).startswith("Row"):
                comments_log.append({
                    "row_ref": row[0],
                    "comment": row[1] if len(row) > 1 else None,
                    "author": row[2] if len(row) > 2 else None,
                    "timestamp": row[3] if len(row) > 3 else None,
                })

    # --- Data completeness: fraction of non-null cells across core columns,
    #     restricted to real (non-excluded, non-on-hold) rows ---
    working_rows = df
    if "not_applicable" in df.columns:
        working_rows = working_rows[~working_rows["not_applicable"].fillna(False)]
    present_core_cols = [c for c in CORE_SIGNAL_COLUMNS if c in df.columns]
    if present_core_cols and not working_rows.empty:
        total_cells = len(working_rows) * len(present_core_cols)
        filled_cells = working_rows[present_core_cols].notna().sum().sum()
        completeness = round(filled_cells / total_cells, 3) if total_cells else 0.0
    else:
        completeness = 0.0

    logger.info(
        "Loaded '%s': %d task rows, %.1f%% data completeness",
        project_name, len(df), completeness * 100,
    )

    return {
        "project_name": project_name,
        "summary": summary,
        "tasks": df,
        "comments_log": comments_log,
        "data_completeness": completeness,
        "source_path": path,
    }


if __name__ == "__main__":
    import sys

    result = load_project_plan(sys.argv[1])
    logger.info("Project: %s", result["project_name"])
    logger.info("Data completeness: %s", result["data_completeness"])
    logger.info("Summary keys: %s", list(result["summary"].keys()))
    logger.info("Task rows: %d", len(result["tasks"]))
    logger.info("Comments log entries: %d", len(result["comments_log"]))
